"""Gerador de relatórios Excel multi-aba segmentados por tipo de anomalia — WP-007."""

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

logger = logging.getLogger(__name__)

_COR_CABECALHO = "1F4E79"
_COR_FONTE_CABECALHO = "FFFFFF"

_RECOMENDACOES: dict[str, str] = {
    "ghost": (
        "Verificar situação no RH e regularizar vínculo CNES ou desligar profissional."
    ),
    "missing": (
        "Cadastrar profissional no CNES local para a competência vigente."
    ),
    "multi_unidades": (
        "Validar se a dupla lotação é estrutural ou erro de cadastro."
    ),
    "acs_tacs": (
        "Revisar lotação: ACS/TACS deve estar em UBS (01), CS II (02) ou equivalente (15)."
    ),
    "ace_tace": (
        "Revisar lotação: ACE/TACE deve estar em CS II (02), CCZ (69), Distrito (22) ou (15)."
    ),
    "rq006_estab_fantasma": (
        "Estabelecimento no CNES local não consta na base nacional. "
        "Verificar se foi desativado ou se há pendência de envio ao DATASUS."
    ),
    "rq007_estab_ausente": (
        "Estabelecimento na base nacional não importado no CNES local. "
        "Executar importação de base atualizada."
    ),
    "rq008_prof_fantasma": (
        "Profissional cadastrado localmente não aparece na base nacional. "
        "Verificar se o envio da competência foi realizado ao DATASUS."
    ),
    "rq009_prof_ausente": (
        "Profissional na base nacional ausente no cadastro local. "
        "Verificar se houve exclusão indevida ou falha de importação."
    ),
    "rq010_divergencia_cbo": (
        "CBO divergente entre local e nacional para o mesmo profissional. "
        "Atualizar o CBO no sistema que estiver desatualizado."
    ),
    "rq011_divergencia_ch": (
        "Carga horária divergente entre local e nacional. "
        "Reconciliar os valores e atualizar o sistema defasado."
    ),
}

_DESCRICOES: dict[str, tuple[str, str]] = {
    "multi_unidades":        ("Profissionais em múltiplas unidades",   "MÉDIA"),
    "acs_tacs":              ("ACS/TACS em unidade incorreta",         "ALTA"),
    "ace_tace":              ("ACE/TACE em unidade incorreta",         "ALTA"),
    "ghost":                 ("Ativo no CNES, ausente/inativo no RH",  "CRÍTICA"),
    "missing":               ("Ativo no RH, ausente no CNES",          "ALTA"),
    "rq006_estab_fantasma":  ("Estabelecimentos fantasma (local)",     "ALTA"),
    "rq007_estab_ausente":   ("Estabelecimentos ausentes no local",    "ALTA"),
    "rq008_prof_fantasma":   ("Profissionais fantasma (CNS)",          "CRÍTICA"),
    "rq009_prof_ausente":    ("Profissionais ausentes local (CNS)",    "ALTA"),
    "rq010_divergencia_cbo": ("Divergência de CBO",                    "MÉDIA"),
    "rq011_divergencia_ch":  ("Divergência de carga horária",          "BAIXA"),
}

_CORES_SEVERIDADE: dict[str, str] = {
    "CRÍTICA": "FF0000",
    "ALTA":    "FF6600",
    "MÉDIA":   "FFD700",
    "BAIXA":   "92D050",
}

_REGRAS_RESUMO: list[tuple[str, str]] = [
    ("RQ-003-B",             "multi_unidades"),
    ("RQ-005 ACS/TACS",      "acs_tacs"),
    ("RQ-005 ACE/TACE",      "ace_tace"),
    ("Ghost Payroll",        "ghost"),
    ("Missing Registration", "missing"),
    ("RQ-006",               "rq006_estab_fantasma"),
    ("RQ-007",               "rq007_estab_ausente"),
    ("RQ-008",               "rq008_prof_fantasma"),
    ("RQ-009",               "rq009_prof_ausente"),
    ("RQ-010",               "rq010_divergencia_cbo"),
    ("RQ-011",               "rq011_divergencia_ch"),
]

_ORDEM_ABAS: list[tuple[str, str]] = [
    ("principal",            "Principal"),
    ("ghost",                "Ghost_Payroll"),
    ("missing",              "Missing_Registro"),
    ("multi_unidades",       "Multi_Unidades"),
    ("acs_tacs",             "ACS_TACS_Incorretos"),
    ("ace_tace",             "ACE_TACE_Incorretos"),
    ("rq006_estab_fantasma", "RQ006 Estab Fantasma"),
    ("rq007_estab_ausente",  "RQ007 Estab Ausente"),
    ("rq008_prof_fantasma",  "RQ008 Prof Fantasma"),
    ("rq009_prof_ausente",   "RQ009 Prof Ausente"),
    ("rq010_divergencia_cbo","RQ010 Diverg CBO"),
    ("rq011_divergencia_ch", "RQ011 Diverg CH"),
]


def gerar_relatorio(
    caminho: Path,
    resultados: dict[str, pd.DataFrame],
    competencia: str = "",
    municipio: str = "Presidente Epitácio/SP",
    metricas: dict | None = None,
) -> None:
    """Gera relatório Excel multi-aba com DataFrames de auditoria segmentados.

    Args:
        caminho: Caminho do arquivo .xlsx de saída.
        resultados: Dicionário chave → DataFrame (aba criada apenas se não vazio).
        competencia: Período no formato AAAA-MM para o cabeçalho do RESUMO.
        municipio: Nome do município exibido no RESUMO.
        metricas: Métricas avançadas opcionais para aba extra.
    """
    caminho.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        _escrever_abas_dados(writer, resultados)
        _gerar_aba_resumo(writer.book, resultados, competencia, municipio)
        if metricas:
            _gerar_aba_metricas(writer.book, metricas)
        total_abas = len(writer.book.sheetnames)

    vinculos = len(resultados.get("principal", pd.DataFrame()))
    logger.info(
        "relatorio gerado caminho=%s abas=%d vinculos=%d",
        caminho, total_abas, vinculos,
    )


def _escrever_abas_dados(writer, resultados: dict[str, pd.DataFrame]) -> None:
    for chave, nome_aba in _ORDEM_ABAS:
        df = resultados.get(chave)
        if df is None or df.empty:
            continue
        if chave == "principal":
            df.to_excel(writer, sheet_name=nome_aba, index=False)
        else:
            _adicionar_recomendacao(df, _RECOMENDACOES[chave]).to_excel(
                writer, sheet_name=nome_aba, index=False
            )
        _formatar_cabecalho(writer.sheets[nome_aba])


def _gerar_aba_resumo(wb, resultados, competencia, municipio) -> None:
    ws = wb.create_sheet("RESUMO", 0)
    _preencher_indicadores(ws, resultados, competencia, municipio)
    _preencher_tabela_anomalias(ws, resultados)
    _ajustar_larguras_resumo(ws)


def _preencher_indicadores(ws, resultados, competencia, municipio) -> None:
    font_bold = Font(bold=True)
    total_vinculos = len(resultados.get("principal", pd.DataFrame()))
    indicadores = [
        ("Município", municipio),
        ("Competência", competencia),
        ("Data de Geração", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total de Vínculos (Local)", total_vinculos),
    ]
    for linha, (rotulo, valor) in enumerate(indicadores, start=1):
        ws.cell(row=linha, column=1, value=rotulo).font = font_bold
        ws.cell(row=linha, column=2, value=valor)


def _preencher_tabela_anomalias(ws, resultados: dict[str, pd.DataFrame]) -> None:
    font_cabecalho = Font(bold=True, color=_COR_FONTE_CABECALHO)
    fill_cabecalho = PatternFill(fill_type="solid", fgColor=_COR_CABECALHO)
    alinhamento = Alignment(horizontal="center", vertical="center")

    linha_ini = 10
    for col, titulo in enumerate(["Regra", "Descrição", "Anomalias", "Severidade"], start=1):
        cell = ws.cell(row=linha_ini, column=col, value=titulo)
        cell.font = font_cabecalho
        cell.fill = fill_cabecalho
        cell.alignment = alinhamento

    total = 0
    for offset, (regra, chave) in enumerate(_REGRAS_RESUMO, start=1):
        contagem = len(resultados.get(chave, pd.DataFrame()))
        total += contagem
        descricao, severidade = _DESCRICOES[chave]
        linha = linha_ini + offset
        ws.cell(row=linha, column=1, value=regra)
        ws.cell(row=linha, column=2, value=descricao)
        ws.cell(row=linha, column=3, value=contagem)
        cell_sev = ws.cell(row=linha, column=4, value=severidade)
        cell_sev.fill = PatternFill(fill_type="solid", fgColor=_CORES_SEVERIDADE[severidade])

    linha_total = linha_ini + len(_REGRAS_RESUMO) + 1
    font_bold = Font(bold=True)
    ws.cell(row=linha_total, column=1, value="TOTAL").font = font_bold
    ws.cell(row=linha_total, column=3, value=total).font = font_bold


def _ajustar_larguras_resumo(ws) -> None:
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12


def _adicionar_recomendacao(df: pd.DataFrame, texto: str) -> pd.DataFrame:
    resultado = df.copy()
    resultado["RECOMENDACAO"] = texto
    return resultado


def _gerar_aba_metricas(wb, metricas: dict) -> None:
    ws = wb.create_sheet("Métricas Avançadas", 1)
    font_bold = Font(bold=True)
    linha = 1

    ws.cell(row=linha, column=1, value="Indicadores Gerais").font = font_bold
    linha += 1
    indicadores = [
        ("Taxa de anomalia", metricas.get("taxa_anomalia_geral")),
        ("P90 CH Total", metricas.get("p90_ch_total")),
        ("Proporção feminina geral", metricas.get("proporcao_feminina_geral")),
        ("Reincidentes", metricas.get("n_reincidentes")),
        ("Taxa de resolução", metricas.get("taxa_resolucao")),
        ("Velocidade de regularização (meses)", metricas.get("velocidade_regularizacao_media")),
    ]
    for rotulo, valor in indicadores:
        ws.cell(row=linha, column=1, value=rotulo)
        ws.cell(row=linha, column=2, value=valor)
        linha += 1
    linha += 1

    linha = _escrever_bloco_json(
        ws, metricas.get("top_glosas_json", "[]"), linha,
        ["CPF", "CNS", "Nome", "Total"],
        "Top Profissionais com Mais Glosas",
    )
    linha = _escrever_bloco_json(
        ws, metricas.get("anomalias_por_cbo_json", "[]"), linha,
        ["CBO", "Descrição", "Total", "Taxa"],
        "Anomalias por CBO",
    )
    _escrever_bloco_json(
        ws, metricas.get("ranking_cnes_json", "[]"), linha,
        ["CNES", "Nome", "Total Anomalias", "Índice de Conformidade"],
        "Ranking CNES por Conformidade",
    )
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30


def _escrever_bloco_json(ws, json_str, linha_ini: int, colunas: list, titulo: str) -> int:
    font_bold = Font(bold=True)
    try:
        items = json.loads(json_str) if json_str else []
    except (json.JSONDecodeError, TypeError):
        items = []
    if not items:
        return linha_ini

    ws.cell(row=linha_ini, column=1, value=titulo).font = font_bold
    for col, cabecalho in enumerate(colunas, start=1):
        ws.cell(row=linha_ini + 1, column=col, value=cabecalho).font = font_bold

    chaves = list(items[0].keys())
    for offset, item in enumerate(items):
        for col, chave in enumerate(chaves, start=1):
            ws.cell(row=linha_ini + 2 + offset, column=col, value=item.get(chave))

    return linha_ini + 1 + 1 + len(items) + 1


def _formatar_cabecalho(ws) -> None:
    fonte = Font(bold=True, color=_COR_FONTE_CABECALHO)
    preenchimento = PatternFill(fill_type="solid", fgColor=_COR_CABECALHO)
    alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for celula in ws[1]:
        celula.font = fonte
        celula.fill = preenchimento
        celula.alignment = alinhamento

    for coluna in ws.columns:
        amostra = list(coluna)[:101]  # cabeçalho + primeiras 100 linhas
        largura = max(
            len(str(celula.value)) if celula.value is not None else 0
            for celula in amostra
        )
        ws.column_dimensions[coluna[0].column_letter].width = min(largura + 4, 60)
