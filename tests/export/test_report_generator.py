"""test_report_generator.py — Testes Unitários do Gerador de Relatórios Excel (WP-007)

Cobertura:
  - Estrutura do .xlsx: abas criadas/omitidas conforme DataFrames não-vazios.
  - Coluna RECOMENDACAO: presente e sem nulos em todas as abas de auditoria.
  - Aba Principal: criada sempre, sem coluna RECOMENDACAO.
  - Aba RESUMO: sempre presente como primeira aba, com indicadores e tabela de anomalias.
  - Abas cross-check (RQ-006 a RQ-011): criadas quando DataFrame não vazio.
  - Qualidade: contagem de linhas preservada, diretório pai criado.
"""

from pathlib import Path

import openpyxl
import pandas as pd

from export.report_generator import gerar_relatorio


def _df_vinculos(n: int = 3) -> pd.DataFrame:
    return pd.DataFrame({
        "CPF": [str(i) * 11 for i in range(n)],
        "NOME_PROFISSIONAL": [f"PROF {i}" for i in range(n)],
        "CBO": ["515105"] * n,
        "COD_CNES": ["0985333"] * n,
        "ESTABELECIMENTO": ["UBS TESTE"] * n,
    })


def _df_ghost(n: int = 2) -> pd.DataFrame:
    df = _df_vinculos(n)
    df["MOTIVO_GHOST"] = ["AUSENTE_NO_RH"] * n
    return df


def _df_missing(n: int = 2) -> pd.DataFrame:
    return pd.DataFrame({
        "CPF": [str(i) * 11 for i in range(n)],
        "NOME": [f"PROF {i}" for i in range(n)],
        "STATUS": ["ATIVO"] * n,
    })


def _df_auditoria(n: int = 2) -> pd.DataFrame:
    return _df_vinculos(n)


def _df_estab_fantasma(n: int = 2) -> pd.DataFrame:
    return pd.DataFrame({
        "CNES": [f"{i:07d}" for i in range(n)],
        "NOME_FANTASIA": [f"ESTAB {i}" for i in range(n)],
        "FONTE": ["LOCAL"] * n,
    })


def _df_divergencia_cbo(n: int = 2) -> pd.DataFrame:
    return pd.DataFrame({
        "CNS": [f"{i:015d}" for i in range(n)],
        "CNES": ["0985333"] * n,
        "CBO_LOCAL": ["515105"] * n,
        "CBO_NACIONAL": ["322255"] * n,
    })


def _df_divergencia_ch(n: int = 2) -> pd.DataFrame:
    return pd.DataFrame({
        "CNS": [f"{i:015d}" for i in range(n)],
        "CNES": ["0985333"] * n,
        "CH_LOCAL": [40] * n,
        "CH_NACIONAL": [20] * n,
        "DELTA_CH": [20] * n,
    })


def _gerar(caminho: Path, **kwargs) -> None:
    defaults = {"principal": _df_vinculos()}
    defaults.update(kwargs)
    gerar_relatorio(caminho, defaults)


def _abas(caminho: Path) -> list[str]:
    wb = openpyxl.load_workbook(caminho, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _buscar_em_coluna(ws, col_busca: int, valor: str, col_retorno: int):
    for row in ws.iter_rows():
        if str(row[col_busca - 1].value) == valor:
            return row[col_retorno - 1].value
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Grupo 1: Estrutura do arquivo
# ─────────────────────────────────────────────────────────────────────────────

class TestEstruturaArquivo:

    def test_cria_arquivo_xlsx(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        assert caminho.exists()
        assert caminho.suffix == ".xlsx"

    def test_arquivo_abre_sem_erro(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        wb = openpyxl.load_workbook(caminho)
        assert wb is not None
        wb.close()

    def test_aba_principal_sempre_criada(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        assert "Principal" in _abas(caminho)

    def test_cria_diretorio_pai_se_inexistente(self, tmp_path: Path):
        caminho = tmp_path / "subdir" / "relatorio.xlsx"
        _gerar(caminho)
        assert caminho.exists()

    def test_aba_ghost_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost())
        assert "Ghost_Payroll" in _abas(caminho)

    def test_aba_ghost_omitida_quando_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=pd.DataFrame())
        assert "Ghost_Payroll" not in _abas(caminho)

    def test_aba_missing_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, missing=_df_missing())
        assert "Missing_Registro" in _abas(caminho)

    def test_aba_missing_omitida_quando_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, missing=pd.DataFrame())
        assert "Missing_Registro" not in _abas(caminho)

    def test_aba_multi_unidades_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, multi_unidades=_df_auditoria())
        assert "Multi_Unidades" in _abas(caminho)

    def test_aba_acs_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, acs_tacs=_df_auditoria())
        assert "ACS_TACS_Incorretos" in _abas(caminho)

    def test_aba_ace_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ace_tace=_df_auditoria())
        assert "ACE_TACE_Incorretos" in _abas(caminho)

    def test_somente_principal_quando_todos_vazios(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        assert _abas(caminho) == ["RESUMO", "Principal"]


# ─────────────────────────────────────────────────────────────────────────────
# Grupo 2: Coluna RECOMENDACAO
# ─────────────────────────────────────────────────────────────────────────────

class TestRecomendacao:

    def test_aba_ghost_tem_coluna_recomendacao(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost())
        df = pd.read_excel(caminho, sheet_name="Ghost_Payroll")
        assert "RECOMENDACAO" in df.columns

    def test_aba_missing_tem_coluna_recomendacao(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, missing=_df_missing())
        df = pd.read_excel(caminho, sheet_name="Missing_Registro")
        assert "RECOMENDACAO" in df.columns

    def test_aba_multi_unidades_tem_coluna_recomendacao(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, multi_unidades=_df_auditoria())
        df = pd.read_excel(caminho, sheet_name="Multi_Unidades")
        assert "RECOMENDACAO" in df.columns

    def test_aba_acs_tem_coluna_recomendacao(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, acs_tacs=_df_auditoria())
        df = pd.read_excel(caminho, sheet_name="ACS_TACS_Incorretos")
        assert "RECOMENDACAO" in df.columns

    def test_aba_ace_tem_coluna_recomendacao(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ace_tace=_df_auditoria())
        df = pd.read_excel(caminho, sheet_name="ACE_TACE_Incorretos")
        assert "RECOMENDACAO" in df.columns

    def test_recomendacao_sem_nulos_em_ghost(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost(5))
        df = pd.read_excel(caminho, sheet_name="Ghost_Payroll")
        assert df["RECOMENDACAO"].isna().sum() == 0

    def test_recomendacao_sem_nulos_em_missing(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, missing=_df_missing(5))
        df = pd.read_excel(caminho, sheet_name="Missing_Registro")
        assert df["RECOMENDACAO"].isna().sum() == 0

    def test_recomendacao_nao_vazia_em_acs(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, acs_tacs=_df_auditoria(3))
        df = pd.read_excel(caminho, sheet_name="ACS_TACS_Incorretos")
        assert (df["RECOMENDACAO"].str.len() > 0).all()

    def test_principal_nao_tem_coluna_recomendacao(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        df = pd.read_excel(caminho, sheet_name="Principal")
        assert "RECOMENDACAO" not in df.columns


# ─────────────────────────────────────────────────────────────────────────────
# Grupo 3: Qualidade dos dados
# ─────────────────────────────────────────────────────────────────────────────

class TestQualidadeDados:

    def test_contagem_linhas_principal_preservada(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, principal=_df_vinculos(7))
        df = pd.read_excel(caminho, sheet_name="Principal")
        assert len(df) == 7

    def test_contagem_linhas_ghost_preservada(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost(4))
        df = pd.read_excel(caminho, sheet_name="Ghost_Payroll")
        assert len(df) == 4

    def test_colunas_originais_preservadas_em_ghost(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost())
        df = pd.read_excel(caminho, sheet_name="Ghost_Payroll")
        assert "CPF" in df.columns
        assert "MOTIVO_GHOST" in df.columns

    def test_colunas_originais_preservadas_em_missing(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, missing=_df_missing())
        df = pd.read_excel(caminho, sheet_name="Missing_Registro")
        assert "CPF" in df.columns
        assert "STATUS" in df.columns


# ─────────────────────────────────────────────────────────────────────────────
# Grupo 4: Aba RESUMO
# ─────────────────────────────────────────────────────────────────────────────

class TestResumo:

    def test_aba_resumo_sempre_presente(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        assert "RESUMO" in _abas(caminho)

    def test_resumo_e_primeira_aba(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho)
        assert _abas(caminho)[0] == "RESUMO"

    def test_resumo_contem_municipio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {"principal": _df_vinculos()}, municipio="Cidade Teste/SP")
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        valores_b = [ws.cell(r, 2).value for r in range(1, 5)]
        assert "Cidade Teste/SP" in valores_b

    def test_resumo_contem_competencia(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {"principal": _df_vinculos()}, competencia="2024-12")
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        valores_b = [ws.cell(r, 2).value for r in range(1, 5)]
        assert "2024-12" in valores_b

    def test_resumo_contem_total_vinculos(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, principal=_df_vinculos(5))
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        valores_b = [ws.cell(r, 2).value for r in range(1, 5)]
        assert 5 in valores_b

    def test_resumo_tabela_anomalias_contagem_correta(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost(3))
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        contagem = _buscar_em_coluna(ws, 1, "Ghost Payroll", 3)
        assert contagem == 3

    def test_resumo_total_anomalias_e_soma(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {
            "ghost": _df_ghost(2),
            "rq006_estab_fantasma": _df_estab_fantasma(3),
        })
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        total = _buscar_em_coluna(ws, 1, "TOTAL", 3)
        assert total == 5

    def test_resumo_severidade_colorida(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, ghost=_df_ghost())
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        tem_fill = False
        for row in ws.iter_rows():
            if row[3].value == "CRÍTICA":
                fill = row[3].fill
                if fill.patternType and fill.patternType != "none":
                    tem_fill = True
                break
        assert tem_fill


# ─────────────────────────────────────────────────────────────────────────────
# Grupo 5: Abas cross-check
# ─────────────────────────────────────────────────────────────────────────────

class TestAbasCrossCheck:

    def test_aba_rq006_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq006_estab_fantasma=_df_estab_fantasma())
        assert "RQ006 Estab Fantasma" in _abas(caminho)

    def test_aba_rq006_omitida_quando_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq006_estab_fantasma=pd.DataFrame())
        assert "RQ006 Estab Fantasma" not in _abas(caminho)

    def test_aba_rq008_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq008_prof_fantasma=_df_estab_fantasma())
        assert "RQ008 Prof Fantasma" in _abas(caminho)

    def test_aba_rq010_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq010_divergencia_cbo=_df_divergencia_cbo())
        assert "RQ010 Diverg CBO" in _abas(caminho)

    def test_aba_rq011_criada_quando_nao_vazio(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq011_divergencia_ch=_df_divergencia_ch())
        assert "RQ011 Diverg CH" in _abas(caminho)

    def test_coluna_recomendacao_presente_em_rq006(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq006_estab_fantasma=_df_estab_fantasma())
        df = pd.read_excel(caminho, sheet_name="RQ006 Estab Fantasma")
        assert "RECOMENDACAO" in df.columns

    def test_coluna_recomendacao_presente_em_rq010(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq010_divergencia_cbo=_df_divergencia_cbo())
        df = pd.read_excel(caminho, sheet_name="RQ010 Diverg CBO")
        assert "RECOMENDACAO" in df.columns

    def test_recomendacao_sem_nulos_em_rq008(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        _gerar(caminho, rq008_prof_fantasma=_df_estab_fantasma(4))
        df = pd.read_excel(caminho, sheet_name="RQ008 Prof Fantasma")
        assert df["RECOMENDACAO"].isna().sum() == 0

    def test_nome_aba_max_31_caracteres(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        resultados = {
            "principal": _df_vinculos(),
            "ghost": _df_ghost(),
            "missing": _df_missing(),
            "multi_unidades": _df_auditoria(),
            "acs_tacs": _df_auditoria(),
            "ace_tace": _df_auditoria(),
            "rq006_estab_fantasma": _df_estab_fantasma(),
            "rq007_estab_ausente": _df_estab_fantasma(),
            "rq008_prof_fantasma": _df_estab_fantasma(),
            "rq009_prof_ausente": _df_estab_fantasma(),
            "rq010_divergencia_cbo": _df_divergencia_cbo(),
            "rq011_divergencia_ch": _df_divergencia_ch(),
        }
        gerar_relatorio(caminho, resultados)
        for nome in _abas(caminho):
            assert len(nome) <= 31, f"Aba '{nome}' excede 31 caracteres"


# ─────────────────────────────────────────────────────────────────────────────
# Grupo 6: Compatibilidade
# ─────────────────────────────────────────────────────────────────────────────

class TestCompatibilidade:

    def test_dict_vazio_gera_apenas_resumo(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {})
        assert _abas(caminho) == ["RESUMO"]

    def test_dict_sem_principal_gera_resumo_com_zero_vinculos(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {"ghost": _df_ghost()})
        wb = openpyxl.load_workbook(caminho)
        ws = wb["RESUMO"]
        valores_b = [ws.cell(r, 2).value for r in range(1, 5)]
        assert 0 in valores_b


def _metricas_sample() -> dict:
    return {
        "taxa_anomalia_geral": 0.15,
        "p90_ch_total": 40.0,
        "proporcao_feminina_geral": 0.6,
        "n_reincidentes": 3,
        "taxa_resolucao": 0.5,
        "velocidade_regularizacao_media": 2.0,
        "top_glosas_json": '[{"cpf":"12345678901","cns":null,"nome":"Ana Silva","total":5}]',
        "anomalias_por_cbo_json": '[{"cbo":"515105","descricao":"ACS","total":3,"taxa":0.1}]',
        "proporcao_feminina_por_cnes_json": '[{"cnes":"1234567","proporcao_f":0.6,"total":10}]',
        "ranking_cnes_json": (
            '[{"cnes":"1234567","nome":"UBS Centro","total_anomalias":2,"indice_conformidade":0.8}]'
        ),
    }


class TestAbaMetricas:

    def test_aba_metricas_ausente_sem_metricas(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {"principal": _df_vinculos()}, metricas=None)
        assert "Métricas Avançadas" not in _abas(caminho)

    def test_aba_metricas_criada_com_metricas(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {"principal": _df_vinculos()}, metricas=_metricas_sample())
        assert "Métricas Avançadas" in _abas(caminho)

    def test_aba_metricas_contem_indicadores(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        gerar_relatorio(caminho, {"principal": _df_vinculos()}, metricas=_metricas_sample())
        wb = openpyxl.load_workbook(caminho)
        ws = wb["Métricas Avançadas"]
        valores = [ws.cell(r, 1).value for r in range(1, 10)]
        assert "Taxa de anomalia" in valores

    def test_aba_metricas_bloco_top_glosas(self, tmp_path: Path):
        caminho = tmp_path / "relatorio.xlsx"
        metricas = _metricas_sample()
        metricas["top_glosas_json"] = '[{"cpf":"123","cns":null,"nome":"Ana","total":5}]'
        gerar_relatorio(caminho, {"principal": _df_vinculos()}, metricas=metricas)
        wb = openpyxl.load_workbook(caminho)
        ws = wb["Métricas Avançadas"]
        todos_valores = [ws.cell(r, c).value for r in range(1, 30) for c in range(1, 5)]
        assert any("Top Profissionais" in str(v) for v in todos_valores if v)
        assert "Ana" in todos_valores
